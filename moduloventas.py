import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import json
from datetime import datetime

# ===============================
# Inicialización del Estado de Sesión
# ===============================

# Inicializar el estado del pedido y el stock si no existen
if 'pedido' not in st.session_state:
    st.session_state.pedido = []

if 'df_productos' not in st.session_state:
    file_path_productos = 'archivo_modificado_productos_20240928_201237.xlsx'  # Archivo de productos
    try:
        st.session_state.df_productos = pd.read_excel(file_path_productos)
    except Exception as e:
        st.error(f"Error al cargar el archivo de productos: {e}")
        st.stop()

if 'df_clientes' not in st.session_state:
    file_path_clientes = 'archivo_modificado_clientes_20240928_200050.xlsx'  # Archivo de clientes
    try:
        st.session_state.df_clientes = pd.read_excel(file_path_clientes)
    except Exception as e:
        st.error(f"Error al cargar el archivo de clientes: {e}")
        st.stop()

# Inicializar 'delete_confirm' como un diccionario si no existe
if 'delete_confirm' not in st.session_state:
    st.session_state.delete_confirm = {}

# Inicializar 'df_equipo' si no existe
if 'df_equipo' not in st.session_state:
    # Definir los miembros del equipo
    data_equipo = {
        'Nombre': [
            'Joni', 'Eduardo', 'Johan', 'Martin',
            'Marian', 'Sofi', 'Valen', 'Emily',
            'Maria-Jose', 'Tu Nombre'
        ],
        'Rol': [
            'Presidente', 'Gerente General', 'Jefe de Depósito', 'Armar Pedidos',
            'Vendedora', 'Vendedora', 'Vendedora', 'Vendedora',
            'Fotógrafa y Catalogador', 'Super Admin'
        ],
        'Departamento': [
            'Dirección', 'Dirección', 'Depósito', 'Depósito',
            'Ventas', 'Ventas', 'Ventas', 'Ventas',
            'Marketing', 'Dirección'
        ],
        'Nivel de Acceso': [
            'Alto', 'Alto', 'Medio', 'Medio',
            'Bajo', 'Bajo', 'Bajo', 'Bajo',
            'Medio', 'Super Admin'
        ]
    }
    st.session_state.df_equipo = pd.DataFrame(data_equipo)

# Inicializar 'usuario' en sesión si no existe
if 'usuario' not in st.session_state:
    st.session_state.usuario = None

# ===============================
# Función para Guardar Pedido en Excel
# ===============================

def guardar_pedido_excel(file_path, order_data):
    try:
        book = load_workbook(file_path)
        if 'Pedidos' in book.sheetnames:
            sheet = book['Pedidos']
        else:
            sheet = book.create_sheet('Pedidos')
            # Escribir encabezados
            sheet.append(['ID Pedido', 'Cliente', 'Vendedor', 'Fecha', 'Hora', 'Items'])
        
        # Generar ID de pedido
        if sheet.max_row == 1:
            id_pedido = 1
        else:
            last_id = sheet['A'][sheet.max_row - 1].value
            id_pedido = last_id + 1 if last_id is not None else 1
        
        # Formatear los ítems como JSON
        items_json = json.dumps(order_data['items'], ensure_ascii=False)
        
        # Agregar nueva fila
        sheet.append([
            id_pedido,
            order_data['cliente'],
            order_data['vendedor'],
            order_data['fecha'],
            order_data['hora'],
            items_json
        ])
        
        # Guardar el libro
        book.save(file_path)
    except Exception as e:
        st.error(f"Error al guardar el pedido: {e}")

# ===============================
# Función de Autenticación con Autocompletado
# ===============================

def login():
    st.sidebar.title("🔒 Iniciar Sesión")
    
    # Campo de texto para ingresar el nombre
    nombre_busqueda = st.sidebar.text_input(
        "Escribe tu nombre",
        placeholder="Comienza a escribir tu nombre...",
        key="nombre_busqueda"
    )
    
    # Filtrar los nombres que contienen la búsqueda (case insensitive)
    if nombre_busqueda:
        opciones_filtradas = st.session_state.df_equipo[
            st.session_state.df_equipo['Nombre'].str.contains(nombre_busqueda, case=False, na=False)
        ]['Nombre'].tolist()
    else:
        opciones_filtradas = st.session_state.df_equipo['Nombre'].tolist()
    
    # Agregar una opción vacía al inicio
    opciones_filtradas = [""] + opciones_filtradas
    
    # Selectbox con las opciones filtradas
    nombre_seleccionado = st.sidebar.selectbox(
        "Selecciona tu nombre",
        opciones_filtradas,
        key="nombre_seleccionado",
        help="Selecciona tu nombre de la lista."
    )
    
    # Si se selecciona un nombre, autenticar al usuario
    if nombre_seleccionado:
        usuario_data = st.session_state.df_equipo[st.session_state.df_equipo['Nombre'] == nombre_seleccionado].iloc[0]
        st.session_state.usuario = {
            'Nombre': usuario_data['Nombre'],
            'Rol': usuario_data['Rol'],
            'Departamento': usuario_data['Departamento'],
            'Nivel de Acceso': usuario_data['Nivel de Acceso']
        }
        st.sidebar.success(f"Bienvenido, {usuario_data['Nombre']} ({usuario_data['Rol']})")
    else:
        st.sidebar.info("Por favor, escribe y selecciona tu nombre para iniciar sesión.")

# ===============================
# Función para Verificar Acceso
# ===============================

def verificar_acceso(nivel_requerido):
    niveles = {
        'Bajo': 1,
        'Medio': 2,
        'Alto': 3,
        'Super Admin': 4
    }
    if st.session_state.usuario:
        usuario_nivel = st.session_state.usuario['Nivel de Acceso']
        if niveles.get(usuario_nivel, 0) >= niveles.get(nivel_requerido, 0):
            return True
    return False

# ===============================
# Configuración de la Página
# ===============================

st.set_page_config(page_title="🛒 Módulo de Ventas", layout="wide")

# Título de la Aplicación
st.title("🐻 Módulo de Ventas 🛒")

# Sidebar para Inicio de Sesión
login()

# Si el usuario no está autenticado, detener la ejecución
if not st.session_state.usuario:
    st.stop()

# Mostrar información del usuario en la parte superior
st.markdown(f"### Usuario: **{st.session_state.usuario['Nombre']}**")
st.markdown(f"### Rol: **{st.session_state.usuario['Rol']}**")
st.markdown("---")

# ===============================
# Navegación entre Módulos
# ===============================

st.sidebar.title("📚 Navegación")
seccion = st.sidebar.radio("Ir a", ["Ventas", "Equipo"])

# ===============================
# Módulo de Ventas
# ===============================

if seccion == "Ventas":
    # Colocamos el buscador de cliente
    col1, col2 = st.columns([2, 1])
    
    with col1:
        cliente_seleccionado = st.selectbox(
            "🔮 Buscar cliente", [""] + st.session_state.df_clientes['Nombre'].unique().tolist(),
            help="Escribí el nombre del cliente o seleccioná uno de la lista."
        )
    
    # Solo mostramos los demás campos si se selecciona un cliente distinto al espacio vacío
    if cliente_seleccionado != "":
        cliente_data = st.session_state.df_clientes[st.session_state.df_clientes['Nombre'] == cliente_seleccionado].iloc[0]
    
        # Mostrar descuento y última compra
        with col1:
            st.write(f"**Descuento:** {cliente_data['Descuento']}%")
            st.write(f"**Última compra:** {cliente_data['Fecha Modificado']}")
    
        # Mostrar vendedor principal
        with col2:
            vendedores = cliente_data['Vendedores'].split(',') if pd.notna(cliente_data['Vendedores']) else ['No asignado']
            vendedor_default = vendedores[0]
            vendedor_seleccionado = st.selectbox("Vendedor", vendedores, index=0)
            st.write(f"**Vendedor Principal:** {vendedor_seleccionado}")
    
        # Sección de productos solo aparece si hay cliente seleccionado
        st.header("📁 Buscador de Productos 🔍")
    
        # Tres columnas: Buscador, precio, y stock con colores
        col_prod1, col_prod2, col_prod3 = st.columns([2, 1, 1])
    
        with col_prod1:
            # Buscador de productos con espacio vacío al inicio
            producto_buscado = st.selectbox(
                "Buscar producto",
                [""] + st.session_state.df_productos['Nombre'].unique().tolist(),
                help="Escribí el nombre del producto o seleccioná uno de la lista."
            )
    
        if producto_buscado:
            producto_data = st.session_state.df_productos[st.session_state.df_productos['Nombre'] == producto_buscado].iloc[0]
    
            with col_prod2:
                # Mostrar precio
                st.write(f"**Precio:** ${producto_data['Precio']}")
    
            with col_prod3:
                # Mostrar stock con colores según la cantidad
                stock = max(0, producto_data['Stock'])  # Nos aseguramos que el stock no sea negativo
                if stock <= 0:
                    color = 'red'
                elif stock < 10:
                    color = 'orange'
                else:
                    color = 'green'
    
                st.markdown(f"<span style='color:{color}'>**Stock disponible:** {stock}</span>", unsafe_allow_html=True)
    
            # Dividimos la sección en dos columnas para mostrar el código y la cantidad en la izquierda, y la imagen a la derecha
            col_izq, col_der = st.columns([2, 1])
    
            with col_izq:
                # Mostrar código del producto
                st.write(f"**Código del producto:** {producto_data['Codigo']}")
    
                # Verificar si la venta está forzada por múltiplos
                if pd.notna(producto_data['forzar multiplos']) and producto_data['forzar multiplos'] > 0:
                    st.warning(f"Este producto tiene venta forzada por {int(producto_data['forzar multiplos'])} unidades.")
                    cantidad = st.number_input(
                        "Cantidad",
                        min_value=int(producto_data['forzar multiplos']),
                        step=int(producto_data['forzar multiplos']),
                        key=f"cantidad_{producto_data['Codigo']}"
                    )
                else:
                    # Campo para seleccionar cantidad si no está forzada la venta por múltiplos
                    if stock > 0:
                        cantidad = st.number_input(
                            "Cantidad",
                            min_value=1,
                            max_value=stock,
                            step=1,
                            key=f"cantidad_{producto_data['Codigo']}"
                        )
                    else:
                        cantidad = 0
                        st.error("No hay stock disponible para este producto.")
    
                # Botón para agregar el producto al pedido, deshabilitado si no hay stock
                boton_agregar_desactivado = stock <= 0  # Deshabilitar el botón si no hay stock
                if st.button("Agregar producto", disabled=boton_agregar_desactivado, key=f"agregar_{producto_data['Codigo']}"):
                    # Verificar si el producto ya está en el pedido
                    existe = any(item['Codigo'] == producto_data['Codigo'] for item in st.session_state.pedido)
                    if existe:
                        st.warning("Este producto ya está en el pedido. Por favor, ajusta la cantidad si es necesario.")
                    else:
                        # Añadir producto al pedido con la cantidad seleccionada
                        producto_agregado = {
                            'Codigo': producto_data['Codigo'],
                            'Nombre': producto_data['Nombre'],
                            'Cantidad': cantidad,
                            'Precio': producto_data['Precio'],
                            'Importe': cantidad * producto_data['Precio']
                        }
                        st.session_state.pedido.append(producto_agregado)
                        # Descontar del stock
                        st.session_state.df_productos.loc[
                            st.session_state.df_productos['Codigo'] == producto_data['Codigo'], 'Stock'
                        ] -= cantidad
                        st.success(f"Se agregó {cantidad} unidad(es) de {producto_data['Nombre']} al pedido.")
    
            with col_der:
                # Mostrar imagen del producto en la columna aparte
                if pd.notna(producto_data['imagen']) and producto_data['imagen'] != '':
                    st.image(producto_data['imagen'], width=200, caption="Imagen del producto")
                else:
                    st.write("No hay imagen disponible.")
    
        # Mostrar el pedido actual
        if st.session_state.pedido:
            st.header("📦 Pedido actual")
    
            # Mostrar la tabla del pedido con la opción de eliminar ítems
            for producto in st.session_state.pedido.copy():  # Use copy to avoid modification during iteration
                codigo = producto['Codigo']
                nombre = producto['Nombre']
                cantidad = producto['Cantidad']
                precio = producto['Precio']
                importe = producto['Importe']
    
                # Crear columnas para mostrar el producto y el botón de eliminar
                col1, col2, col3, col4, col5, col6 = st.columns([1, 2, 1, 1, 1, 1])
                col1.write(codigo)
                col2.write(nombre)
                col3.write(cantidad)
                col4.write(f"${precio}")
                col5.write(f"${importe}")
    
                # Verificar si este producto está pendiente de eliminación
                if codigo in st.session_state.delete_confirm:
                    # Mostrar botón de "Sí" y "No" en rojo
                    with col6:
                        # Botón "Sí" para confirmar eliminación
                        if st.button("Sí", key=f"confirmar_si_{codigo}"):
                            # Eliminar el ítem del pedido
                            index = next((i for i, item in enumerate(st.session_state.pedido) if item['Codigo'] == codigo), None)
                            if index is not None:
                                producto_eliminado = st.session_state.pedido.pop(index)
                                # Reponer el stock
                                st.session_state.df_productos.loc[
                                    st.session_state.df_productos['Codigo'] == producto_eliminado['Codigo'], 'Stock'
                                ] += producto_eliminado['Cantidad']
                            # Remover del diccionario de confirmaciones
                            del st.session_state.delete_confirm[codigo]
    
                        # Botón "No" para cancelar eliminación
                        if st.button("No", key=f"confirmar_no_{codigo}"):
                            # Cancelar la eliminación
                            del st.session_state.delete_confirm[codigo]
                else:
                    # Mostrar el botón de eliminar normal
                    with col6:
                        if st.button('🗑️', key=f"eliminar_{codigo}"):
                            # Marcar este ítem para eliminación
                            st.session_state.delete_confirm[codigo] = True
    
            # Calcular totales
            pedido_df = pd.DataFrame(st.session_state.pedido)
            total_items = pedido_df['Cantidad'].sum() if not pedido_df.empty else 0
            total_monto = pedido_df['Importe'].sum() if not pedido_df.empty else 0.0
    
            # Mostrar total de ítems y total del pedido en una sola fila
            col_items, col_total = st.columns([1, 1])
    
            with col_items:
                st.write(f"**Total de ítems:** {total_items}")
    
            with col_total:
                # Mostrar total del pedido al lado de total de ítems
                st.write(f"<h4 style='text-align:right;'>Total del pedido: ${total_monto:,.2f}</h4>", unsafe_allow_html=True)
    
            # Centrar el botón de guardar pedido
            col_guardar, _ = st.columns([2, 3])
            with col_guardar:
                if st.button("Guardar Pedido"):
                    if not st.session_state.pedido:
                        st.warning("No hay ítems en el pedido para guardar.")
                    else:
                        # Obtener fecha y hora actuales
                        now = datetime.now()
                        fecha_actual = now.strftime("%Y-%m-%d")
                        hora_actual = now.strftime("%H:%M:%S")
    
                        # Preparar datos del pedido
                        order_data = {
                            'cliente': cliente_seleccionado,
                            'vendedor': vendedor_seleccionado,
                            'fecha': fecha_actual,
                            'hora': hora_actual,
                            'items': st.session_state.pedido
                        }
    
                        # Guardar el pedido en la hoja 'Pedidos'
                        guardar_pedido_excel(file_path_productos, order_data)
    
                        # Confirmar al usuario
                        st.success("Pedido guardado exitosamente.", icon="✅")
    
                        # Limpiar el pedido después de guardarlo
                        st.session_state.pedido = []
                        st.session_state.delete_confirm = {}
    
                        # Guardar los cambios en el stock de productos
                        try:
                            with pd.ExcelWriter(file_path_productos, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                                st.session_state.df_productos.to_excel(writer, sheet_name='Hoja1', index=False)
                        except Exception as e:
                            st.error(f"Error al actualizar el stock en el archivo de productos: {e}")

# ===============================
# Módulo de Equipo
# ===============================

elif seccion == "Equipo":
    # Verificar el nivel de acceso necesario para ver el módulo de equipo
    if not verificar_acceso('Medio'):
        st.error("No tienes permisos para acceder a esta sección.")
        st.stop()
    
    st.header("👥 Equipo de Trabajo")
    
    # Mostrar la tabla del equipo
    st.dataframe(st.session_state.df_equipo, use_container_width=True)
    
    st.markdown("---")
    
    # Opciones de gestión solo para Super Admin
    if st.session_state.usuario['Nivel de Acceso'] == 'Super Admin':
        st.subheader("🔧 Gestionar Equipo")
        
        # Formulario para agregar un nuevo miembro al equipo
        with st.expander("Agregar Nuevo Miembro"):
            with st.form("form_agregar"):
                nombre = st.text_input("Nombre")
                rol = st.selectbox("Rol", [
                    'Presidente', 'Gerente General', 'Jefe de Depósito', 'Armar Pedidos',
                    'Vendedora', 'Fotógrafa y Catalogador', 'Super Admin'
                ])
                departamento = st.selectbox("Departamento", [
                    'Dirección', 'Depósito', 'Ventas', 'Marketing'
                ])
                nivel_acceso = st.selectbox("Nivel de Acceso", [
                    'Bajo', 'Medio', 'Alto', 'Super Admin'
                ])
                submit = st.form_submit_button("Agregar")
                
                if submit:
                    if nombre.strip() == "":
                        st.error("El nombre no puede estar vacío.")
                    elif nombre.strip() in st.session_state.df_equipo['Nombre'].values:
                        st.error("El nombre ya existe en el equipo.")
                    else:
                        nuevo_miembro = {
                            'Nombre': nombre.strip(),
                            'Rol': rol,
                            'Departamento': departamento,
                            'Nivel de Acceso': nivel_acceso
                        }
                        st.session_state.df_equipo = st.session_state.df_equipo.append(nuevo_miembro, ignore_index=True)
                        st.success(f"Miembro {nombre} agregado exitosamente.")
        
        st.markdown("---")
        
        # Formulario para eliminar un miembro del equipo
        with st.expander("Eliminar Miembro"):
            with st.form("form_eliminar"):
                nombre_eliminar = st.selectbox(
                    "Selecciona el nombre a eliminar",
                    st.session_state.df_equipo['Nombre'].unique().tolist()
                )
                submit_eliminar = st.form_submit_button("Eliminar")
                
                if submit_eliminar:
                    if nombre_eliminar in st.session_state.df_equipo['Nombre'].values:
                        if nombre_eliminar == st.session_state.usuario['Nombre']:
                            st.error("No puedes eliminarte a ti mismo.")
                        else:
                            st.session_state.df_equipo = st.session_state.df_equipo[st.session_state.df_equipo['Nombre'] != nombre_eliminar]
                            st.success(f"Miembro {nombre_eliminar} eliminado exitosamente.")
                    else:
                        st.error("El nombre seleccionado no existe.")

    # Mostrar detalles adicionales según el rol del usuario
    st.markdown("---")
    st.subheader("🔍 Detalles del Equipo")
    seleccionado = st.selectbox("Selecciona un miembro del equipo para ver detalles", [""] + st.session_state.df_equipo['Nombre'].tolist())
    
    if seleccionado:
        miembro = st.session_state.df_equipo[st.session_state.df_equipo['Nombre'] == seleccionado].iloc[0]
        st.write(f"**Nombre:** {miembro['Nombre']}")
        st.write(f"**Rol:** {miembro['Rol']}")
        st.write(f"**Departamento:** {miembro['Departamento']}")
        st.write(f"**Nivel de Acceso:** {miembro['Nivel de Acceso']}")
        # Aquí puedes agregar más detalles según necesites

# ===============================
# Opciones de Logout
# ===============================

st.sidebar.markdown("---")
if st.sidebar.button("Cerrar Sesión"):
    st.session_state.usuario = None
    st.experimental_rerun()
